#!/usr/bin/env python3
"""
Turns the Ethiopian material price workbook into a seed migration.

    python3 scripts/build-price-seed.py \
        supabase/seed/Ethiopia_Construction_Materials_Price_Template.xlsx \
        supabase/migrations/0042_material_price_seed.sql

Run once per workbook revision. The output is committed, so applying the seed
never depends on having Python, openpyxl, or the workbook to hand — the owner
runs migrations and the prices are there.

## Why the two sheets land at different statuses

The workbook keeps two lists and says so in its own README: `Price Data` is a
teaching baseline for all 445 materials, and `Current Sourced Prices` is a short
list of figures actually read off a public listing or a published index, each
with a named source and a date.

Those are not the same kind of fact, and the schema has a status for each. A
Jiji listing from August is weak evidence, but it is evidence; a planning
baseline is not evidence at all. Flattening both to `educational_estimate`
would throw away the distinction the workbook took the trouble to draw, and
`web_sourced` outranks `educational_estimate` in the resolver precisely so the
better one wins.

So:

  * `Price Data` rows with no source     -> educational_estimate
  * `Price Data` rows carrying a source  -> web_sourced
  * `Current Sourced Prices`             -> web_sourced

Rows in the third list that exactly duplicate one already in the second are
dropped. They are the same observation written down twice, and importing both
would put one listing into the market average twice over.
"""

from __future__ import annotations

import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# The workbook has no category column on the sourced sheet. Materials that
# appear on both sheets take their category from the main one; these are the
# few that do not appear there under the same name.
FALLBACK_CATEGORY = {
    "gypsum powder": ("Gypsum & Ceiling", "Gypsum Boards"),
    "cement": ("Cement & Concrete", "Cement"),
    "paint": ("Finishes", "Paint"),
}

# The seed's city column. The workbook writes "Addis Ababa (educational
# baseline)" on every baseline row, which says in the city field something the
# status field already says better. Two places to read the same caveat is two
# places for it to go out of date, so the caveat lives in the status and the
# city says where.
BASELINE_CITY = "Addis Ababa (educational baseline)"


def sql_text(value) -> str:
    """A SQL literal. Single quotes doubled; nothing else is interpolated."""
    if value is None:
        return "null"
    text = str(value).strip()
    if not text:
        return "null"
    return "'" + text.replace("'", "''") + "'"


def sql_date(value) -> str:
    if isinstance(value, datetime):
        value = value.date()
    if not isinstance(value, date):
        return "current_date"
    return f"date '{value.isoformat()}'"


def sql_number(value) -> str:
    if value is None:
        return "null"
    return f"{float(value):.2f}"


def vat_status(value) -> str:
    """The workbook writes free text; the column is an enum."""
    text = (str(value) if value is not None else "").strip().lower()
    if text in {"inclusive", "incl", "vat included", "included"}:
        return "'inclusive'"
    if text in {"exclusive", "excl", "vat excluded", "excluded"}:
        return "'exclusive'"
    if text in {"exempt", "exempted"}:
        return "'exempt'"
    return "'unknown'"


def main(workbook_path: str, out_path: str) -> None:
    book = openpyxl.load_workbook(workbook_path, data_only=True)

    price_data = [
        row
        for row in book["Price Data"].iter_rows(min_row=2, values_only=True)
        if row[2]
    ]
    sourced = [
        row
        for row in book["Current Sourced Prices"].iter_rows(min_row=2, values_only=True)
        if row[0]
    ]

    categories: dict[str, tuple[str, str]] = {}
    for row in price_data:
        categories.setdefault(str(row[2]).strip().lower(), (row[0], row[1]))

    records: list[dict] = []
    # Exact observations already recorded, so the sourced sheet does not add a
    # second copy of a listing the main sheet already carries.
    seen: set[tuple] = set()

    for row in price_data:
        (
            category, subcategory, material, specification, unit, brand,
            city, price, vat, supplier, price_date, source, _verified,
            notes, _status,
        ) = row[:15]

        if price is None:
            continue

        # A row that names where it came from is evidence, whatever the
        # workbook's blanket status column says.
        status = "web_sourced" if source else "educational_estimate"

        records.append({
            "category": category,
            "subcategory": subcategory,
            "material": material,
            "specification": specification,
            "unit": unit,
            "brand": brand,
            "city": "Addis Ababa" if city == BASELINE_CITY else city,
            "price": price,
            "vat": vat,
            "supplier": supplier,
            "date": price_date,
            "source": source,
            "notes": notes,
            "status": status,
        })

        if source:
            seen.add((
                str(material).strip().lower(),
                str(specification or "").strip().lower(),
                str(brand or "").strip().lower(),
                str(unit).strip().lower(),
                float(price),
            ))

    duplicates = 0
    for row in sourced:
        material, specification, brand, unit, price, city, price_date, source, confidence, notes = row[:10]
        if price is None:
            continue

        key = (
            str(material).strip().lower(),
            str(specification or "").strip().lower(),
            str(brand or "").strip().lower(),
            str(unit).strip().lower(),
            float(price),
        )
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)

        category, subcategory = categories.get(
            str(material).strip().lower(),
            FALLBACK_CATEGORY.get(str(material).strip().lower(), ("Construction Consumables", None)),
        )

        detail = f"Confidence: {confidence}." if confidence else ""
        records.append({
            "category": category,
            "subcategory": subcategory,
            "material": material,
            "specification": specification,
            "unit": unit,
            "brand": brand,
            "city": city,
            "price": price,
            "vat": None,
            "supplier": None,
            "date": price_date,
            "source": source,
            "notes": f"{notes or ''} {detail}".strip() or None,
            "status": "web_sourced",
        })

    baseline = sum(1 for r in records if r["status"] == "educational_estimate")
    web = sum(1 for r in records if r["status"] == "web_sourced")

    lines: list[str] = []
    lines.append(f"""-- Initial Ethiopian construction material prices.
--
-- Generated from the price workbook by scripts/build-price-seed.py. Do not edit
-- by hand: regenerate it, so the file and the workbook cannot drift apart.
--
--   {len(records)} prices — {baseline} planning baselines, {web} source-backed observations.
--   {duplicates} row(s) from the sourced sheet were dropped as exact duplicates of
--   an observation already present.
--
-- Every figure here is a starting point, not a quotation. The baselines are
-- teaching numbers and the sourced ones are public listings — neither has been
-- confirmed with a supplier, which is what `admin_verified` is for. The
-- resolver ranks both below anything an administrator has stood behind, and the
-- exchange labels them on screen.
--
-- Idempotent: re-running it does not double the book. Rows are matched on the
-- natural key a person would use to say "that is the same price" — material,
-- specification, brand, unit, city, date and figure.
--
-- Safe to paste whole into the Supabase SQL editor. It is one guard plus one
-- insert, with no temporary tables and no explicit transaction, so nothing here
-- depends on two statements landing in the same session.
--
-- Additive. Run after 0041. It creates no tables and alters none: the only
-- thing it writes is rows into public.material_prices.

do $$
begin
  if to_regclass('public.material_prices') is null then
    raise exception using
      message = 'Material price seed: the price book does not exist.',
      hint = 'Run migration 0041 first.';
  end if;
end $$;

-- One statement, start to finish.
--
-- An earlier version staged the rows in a `create temporary table seed_prices
-- (...) on commit drop`, filled it, then read from it — three statements that
-- had to share one session *and* one transaction. That holds in psql and falls
-- over in the Supabase SQL editor, which wraps a submission in its own
-- transaction, lets you run a selection rather than the whole file, and pools
-- connections. Any of those ends the transaction between the create and the
-- read, `on commit drop` takes the table with it, and the next line fails with
-- `42P01: relation "seed_prices" does not exist`.
--
-- The staging table was never anything but scaffolding, so it is gone. The rows
-- live in a CTE inside the insert that consumes them, which cannot be separated
-- from its own data by anything.
with raw (
  category, subcategory, material, specification, unit, brand, city_region,
  price_etb, vat_status, supplier, price_date, source, notes, data_status
) as (
  values""")

    values = []
    for record in records:
        values.append(
            "  ("
            + ", ".join([
                sql_text(record["category"]),
                sql_text(record["subcategory"]),
                sql_text(record["material"]),
                sql_text(record["specification"]),
                sql_text(record["unit"]),
                sql_text(record["brand"]),
                sql_text(record["city"]),
                sql_number(record["price"]),
                vat_status(record["vat"]),
                sql_text(record["supplier"]),
                sql_date(record["date"]),
                sql_text(record["source"]),
                sql_text(record["notes"]),
                f"'{record['status']}'",
            ])
            + ")"
        )

    lines.append(",\n".join(values))

    lines.append("""),

-- Types, stated once.
--
-- A bare `values` list infers each column from its literals, so a column that
-- is null on every row arrives as text and a status arrives as `unknown`.
-- Comparing those against numeric, date and enum columns in the `not exists`
-- below is an error, not a silent coercion — so every column is cast here and
-- the comparison downstream is like for like.
seed_prices as (
  select
    category::text          as category,
    subcategory::text       as subcategory,
    material::text          as material,
    specification::text     as specification,
    unit::text              as unit,
    brand::text             as brand,
    city_region::text       as city_region,
    price_etb::numeric(14, 2) as price_etb,
    vat_status::public.price_vat_status   as vat_status,
    supplier::text          as supplier,
    price_date::date        as price_date,
    source::text            as source,
    notes::text             as notes,
    data_status::public.price_data_status as data_status
  from raw
),

-- One row per observation, before anything is written.
--
-- `not exists` below compares against what is already in the table; it cannot
-- see the rest of this same batch. Two rows identical on the natural key would
-- therefore both pass it and both be inserted. `distinct on` settles that here.
--
-- Note what is *not* in the key: two rows for the same material at different
-- prices, or on different dates, are two observations and both are kept. That
-- is the whole point of an append-only price book, and collapsing them would
-- destroy the history the charts are drawn from.
deduped as (
  select distinct on (
    material, unit, price_etb, price_date, city_region, specification, brand
  ) *
  from seed_prices
  order by material, unit, price_etb, price_date, city_region, specification,
           brand, data_status desc
)

-- Only what is not already there. `not exists` rather than `on conflict`
-- because the natural key is not a unique constraint and must not become one:
-- the same material at the same price on a *different* date is a second
-- observation, and the history depends on being able to hold both.
insert into public.material_prices
  (category, subcategory, material, specification, unit, brand, city_region,
   price_etb, vat_status, supplier, price_date, source, notes, data_status)
select
  s.category, s.subcategory, s.material, s.specification, s.unit, s.brand,
  s.city_region, s.price_etb, s.vat_status, s.supplier, s.price_date, s.source,
  s.notes, s.data_status
from deduped s
where not exists (
  select 1
  from public.material_prices m
  where m.material = s.material
    and m.unit = s.unit
    and m.price_etb = s.price_etb
    and m.price_date = s.price_date
    and m.city_region = s.city_region
    and m.specification is not distinct from s.specification
    and m.brand is not distinct from s.brand
);
""")

    with open(out_path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))

    print(f"{len(records)} prices -> {out_path}")
    print(f"   {baseline} educational_estimate, {web} web_sourced")
    print(f"   {duplicates} duplicate observation(s) dropped")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
